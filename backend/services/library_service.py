import logging
from typing import List, Dict, Optional
from datetime import datetime
from database.db import SessionLocal
from database.models import SavedArticle
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

logger = logging.getLogger(__name__)


class LibraryService:
    """
    Service for managing saved articles library (roczna biblioteka)
    """
    
    async def save_article(self, title: str, url: str, source: str, 
                          description: str, category: str,
                          user_note: str = "", tags: str = "") -> SavedArticle:
        """Save article to library"""
        try:
            db = SessionLocal()
            
            article = SavedArticle(
                title=title,
                url=url,
                source=source,
                description=description,
                category=category,
                user_note=user_note,
                tags=tags
            )
            
            db.add(article)
            db.commit()
            db.refresh(article)
            db.close()
            
            logger.info(f"✅ Article saved: {article.id}")
            return article
        
        except Exception as e:
            logger.error(f"❌ Save error: {str(e)}")
            raise
    
    async def get_articles(self, month: Optional[int] = None,
                          tag: Optional[str] = None) -> List[Dict]:
        """Get articles with optional filters"""
        try:
            db = SessionLocal()
            
            query = db.query(SavedArticle)
            
            # Filter by month
            if month:
                query = query.filter(
                    func.extract('month', SavedArticle.saved_at) == month
                )
            
            # Filter by tag
            if tag:
                query = query.filter(
                    SavedArticle.tags.contains(tag)
                )
            
            articles = query.order_by(SavedArticle.saved_at.desc()).all()
            
            result = []
            for article in articles:
                result.append({
                    "id": article.id,
                    "title": article.title,
                    "url": article.url,
                    "source": article.source,
                    "description": article.description,
                    "category": article.category,
                    "user_note": article.user_note or "",
                    "ai_generated_note": article.ai_generated_note or "",
                    "tags": article.tags or "",
                    "saved_at": article.saved_at.isoformat()
                })
            
            db.close()
            logger.info(f"✅ Retrieved {len(result)} articles")
            return result
        
        except Exception as e:
            logger.error(f"❌ Get articles error: {str(e)}")
            raise
    
    async def update_article(self, article_id: str, user_note: Optional[str] = None,
                            tags: Optional[str] = None,
                            ai_generated_note: Optional[str] = None) -> SavedArticle:
        """Update article"""
        try:
            db = SessionLocal()
            
            article = db.query(SavedArticle).filter(
                SavedArticle.id == article_id
            ).first()
            
            if not article:
                raise ValueError(f"Article {article_id} not found")
            
            if user_note is not None:
                article.user_note = user_note
            if tags is not None:
                article.tags = tags
            if ai_generated_note is not None:
                article.ai_generated_note = ai_generated_note
            
            db.commit()
            db.refresh(article)
            db.close()
            
            logger.info(f"✅ Article updated: {article_id}")
            return article
        
        except Exception as e:
            logger.error(f"❌ Update error: {str(e)}")
            raise
    
    async def delete_article(self, article_id: str) -> bool:
        """Delete article"""
        try:
            db = SessionLocal()
            
            article = db.query(SavedArticle).filter(
                SavedArticle.id == article_id
            ).first()
            
            if not article:
                raise ValueError(f"Article {article_id} not found")
            
            db.delete(article)
            db.commit()
            db.close()
            
            logger.info(f"✅ Article deleted: {article_id}")
            return True
        
        except Exception as e:
            logger.error(f"❌ Delete error: {str(e)}")
            raise
    
    async def export_excel(self) -> str:
        """Export all articles to Excel"""
        try:
            db = SessionLocal()
            
            articles = db.query(SavedArticle).order_by(
                SavedArticle.saved_at.desc()
            ).all()
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Biblioteka MOBGLO"
            
            # Headers
            headers = ["Data", "Artykuł", "Źródło", "Kategoria", "Notatka", "Notatka AI", "Tagi"]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Add data
            for article in articles:
                ws.append([
                    article.saved_at.strftime("%d-%m-%Y"),
                    article.title,
                    article.source,
                    article.category,
                    article.user_note or "",
                    article.ai_generated_note or "",
                    article.tags or ""
                ])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 20
            
            # Save file
            filename = f"MOBGLO-biblioteka-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            filepath = os.path.join("/tmp", filename)
            wb.save(filepath)
            
            db.close()
            logger.info(f"✅ Excel exported: {filename}")
            return filepath
        
        except Exception as e:
            logger.error(f"❌ Export error: {str(e)}")
            raise
